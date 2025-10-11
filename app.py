import os
import json
import cv2
import numpy as np
import re
import pandas as pd
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
from werkzeug.utils import secure_filename
import google.generativeai as genai
from flask_cors import CORS
import base64
from PIL import Image
import io
import traceback
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# Flask Setup
# =========================
app = Flask(__name__, static_folder='static', template_folder='templates')
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["DATA_FOLDER"] = "data"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB max file size
app.static_folder = 'static'
app.static_url_path = '/static'

# Create necessary directories
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["DATA_FOLDER"], exist_ok=True)
os.makedirs("static", exist_ok=True)
os.makedirs("templates", exist_ok=True)

# Enable CORS for frontend integration
CORS(app, origins=['*'])

# =========================
# Configure Gemini AI
# =========================
API_KEY = "YOUR GEMINI API KEY"  # Replace with your actual API key
genai.configure(api_key=API_KEY)

try:
    model = genai.GenerativeModel("gemini-2.0-flash-exp")
    print("✅ Gemini AI model initialized successfully")
except Exception as e:
    print(f"❌ Error initializing Gemini AI: {e}")
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        print("✅ Fallback to Gemini 1.5 Flash")
    except Exception as e2:
        print(f"❌ Fallback model failed: {e2}")
        model = None

# =========================
# Data Storage Files
# =========================
ANSWER_KEYS_FILE = os.path.join(app.config["DATA_FOLDER"], "answer_keys.json")
CONSOLIDATED_REPORTS_FILE = os.path.join(app.config["DATA_FOLDER"], "consolidated_reports.json")
SETTINGS_FILE = os.path.join(app.config["DATA_FOLDER"], "settings.json")
LOGS_FILE = os.path.join(app.config["DATA_FOLDER"], "system_logs.json")

# System start time for uptime calculation
SYSTEM_START_TIME = datetime.now()

# =========================
# Helper Functions
# =========================
def load_json_data(filepath, default=None):
    """Load JSON data from file"""
    if default is None:
        default = []
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return default

def save_json_data(filepath, data):
    """Save JSON data to file"""
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving {filepath}: {e}")
        return False

def log_system_event(level, message):
    """Log system events"""
    logs = load_json_data(LOGS_FILE, [])
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "level": level,
        "message": message
    }
    logs.insert(0, log_entry)  # Add to beginning
    # Keep only last 1000 logs
    logs = logs[:1000]
    save_json_data(LOGS_FILE, logs)
    print(f"[{level.upper()}] {message}")

def get_system_uptime():
    """Calculate system uptime"""
    uptime_delta = datetime.now() - SYSTEM_START_TIME
    hours = int(uptime_delta.total_seconds() // 3600)
    minutes = int((uptime_delta.total_seconds() % 3600) // 60)
    return f"{hours}h {minutes}m"

def allowed_file(filename):
    """Check if file extension is allowed"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'bmp', 'tiff'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_image_with_ai(image_path, questions_count):
    """Process image with Gemini AI to extract answers"""
    if not model:
        raise Exception("AI model not available")
    
    try:
        # Load image
        with open(image_path, 'rb') as f:
            image_data = f.read()
        
        # Create the prompt for answer detection
        prompt = f"""
        Analyze this handwritten multiple choice answer sheet image. 
        This page contains {questions_count} questions.
        
        For each question, identify which option (A, B, C, or D) is marked/circled/filled.
        Look for:
        - Darkened circles or bubbles
        - Check marks
        - X marks
        - Any clear indication of selection
        
        Return ONLY the selected answers in order, one letter per line.
        If a question is unclear or unmarked, respond with 'X'.
        
        Example format:
        A
        B
        C
        D
        A
        
        Analyze the image now:
        """
        
        # Prepare image for API
        image = Image.open(io.BytesIO(image_data))
        
        # Convert to RGB if needed
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Generate content with AI
        response = model.generate_content(
            [prompt, image],
            generation_config={
                "temperature": 0.1,
                "top_p": 0.8,
                "top_k": 40,
                "max_output_tokens": 1000,
            }
        )
        
        if not response or not response.text:
            raise Exception("Empty response from AI")
        
        # Extract answers from response
        response_text = response.text.strip().upper()
        print(f"AI Response: {response_text}")
        
        # Parse the response to extract answers
        answers = []
        lines = response_text.split('\n')
        
        for line in lines:
            line = line.strip()
            # Look for single letters A, B, C, D, or X
            if len(line) == 1 and line in ['A', 'B', 'C', 'D', 'X']:
                answers.append(line)
            elif len(line) > 1:
                # Try to extract letters from longer lines
                matches = re.findall(r'\b[ABCDX]\b', line)
                answers.extend(matches)
        
        # Ensure we have the right number of answers
        while len(answers) < questions_count:
            answers.append('X')  # Fill missing with X
        
        # Truncate if too many
        answers = answers[:questions_count]
        
        print(f"Extracted answers: {answers}")
        return answers
        
    except Exception as e:
        print(f"AI processing error: {e}")
        # Return default answers if AI fails
        return ['X'] * questions_count

def create_enhanced_excel_report(consolidated_data):
    """Create enhanced Excel report with multiple sheets and formatting"""
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Summary Sheet
    summary_sheet = workbook.create_sheet("Test Summary", 0)
    summary_data = [
        ["Test Information", ""],
        ["Test ID", consolidated_data.get('testId', '')],
        ["Test Title", consolidated_data.get('testTitle', '')],
        ["Staff Name", consolidated_data.get('staffName', '')],
        ["Total Questions", consolidated_data.get('totalQuestions', 0)],
        ["Marks Per Question", consolidated_data.get('marksPerQuestion', 0)],
        ["Passing Marks (%)", consolidated_data.get('passingMarks', 0)],
        ["", ""],
        ["Statistics", ""],
        ["Total Students", consolidated_data.get('totalStudents', 0)],
        ["Students Passed", consolidated_data.get('passedCount', 0)],
        ["Students Failed", consolidated_data.get('failedCount', 0)],
        ["Pass Rate (%)", consolidated_data.get('passRate', 0)],
        ["Average Score (%)", consolidated_data.get('averageScore', 0)],
        ["Highest Score (%)", consolidated_data.get('highestScore', 0)],
        ["Lowest Score (%)", consolidated_data.get('lowestScore', 0)],
        ["", ""],
        ["Report Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary_sheet.cell(row=row_idx, column=2, value=value)
    
    # Auto-adjust column widths
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Student Results Sheet
    results_sheet = workbook.create_sheet("Student Results", 1)
    
    # Headers
    headers = [
        'S.No', 'Roll Number', 'Student Name', 'Total Questions', 
        'Correct Answers', 'Wrong Answers', 'Obtained Marks', 
        'Total Marks', 'Percentage (%)', 'Grade', 'Status', 
        'Evaluated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = results_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Student data
    student_results = consolidated_data.get('studentResults', [])
    for row_idx, student in enumerate(student_results, 2):
        summary = student.get('summary', {})
        evaluated_at = datetime.fromisoformat(
            student.get('evaluatedAt', '').replace('Z', '+00:00')
        ).strftime('%Y-%m-%d %H:%M')
        
        row_data = [
            row_idx - 1,  # S.No
            student.get('rollNumber', ''),
            student.get('studentName', ''),
            summary.get('totalQuestions', 0),
            summary.get('correctAnswers', 0),
            summary.get('wrongAnswers', 0),
            summary.get('obtainedMarks', 0),
            summary.get('totalMarks', 0),
            summary.get('percentage', 0),
            summary.get('grade', ''),
            summary.get('status', ''),
            evaluated_at
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = results_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Color code status
            if col_idx == 11:  # Status column
                if value == 'PASS':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    for column in results_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        results_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # 3. Detailed Analysis Sheet
    analysis_sheet = workbook.create_sheet("Detailed Analysis", 2)
    
    # Question-wise analysis
    if student_results:
        total_questions = consolidated_data.get('totalQuestions', 0)
        analysis_data = []
        
        for q_num in range(1, total_questions + 1):
            correct_count = 0
            total_count = len(student_results)
            
            for student in student_results:
                results = student.get('results', [])
                if len(results) >= q_num:
                    if results[q_num - 1].get('isCorrect', False):
                        correct_count += 1
            
            accuracy = (correct_count / total_count * 100) if total_count > 0 else 0
            analysis_data.append([q_num, correct_count, total_count, f"{accuracy:.1f}%"])
        
        # Headers for analysis
        analysis_headers = ['Question No.', 'Correct Answers', 'Total Students', 'Accuracy (%)']
        for col_idx, header in enumerate(analysis_headers, 1):
            cell = analysis_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Analysis data
        for row_idx, data in enumerate(analysis_data, 2):
            for col_idx, value in enumerate(data, 1):
                cell = analysis_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                # Color code accuracy
                if col_idx == 4:  # Accuracy column
                    accuracy_val = float(value.replace('%', ''))
                    if accuracy_val >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif accuracy_val >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in analysis_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            analysis_sheet.column_dimensions[column_letter].width = adjusted_width
    
    return workbook

# =========================
# Routes
# =========================

@app.route('/')
def index():
    """Serve the main application"""
    try:
        # Check if index.html exists in templates
        template_path = os.path.join('templates', 'index.html')
        if os.path.exists(template_path):
            return render_template('index.html')
        else:
            # Return a basic HTML page if template doesn't exist
            return """
            <!DOCTYPE html>
            <html>
            <head>
                <title>Smart Exam Evaluator</title>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
            </head>
            <body>
                <h1>Smart Exam Evaluator Backend</h1>
                <p>Backend server is running. Please ensure your frontend is properly set up.</p>
                <p>API Base URL: <code>/api</code></p>
                <h3>Available Endpoints:</h3>
                <ul>
                    <li>POST /api/test/setup - Setup new test</li>
                    <li>GET /api/test/active - Get active test</li>
                    <li>POST /api/student/evaluate - Evaluate student</li>
                    <li>GET /api/reports/consolidated - Get consolidated reports</li>
                    <li>POST /api/admin/login - Admin login</li>
                    <li>GET /api/dashboard/stats - Dashboard statistics</li>
                </ul>
            </body>
            </html>
            """
    except Exception as e:
        return f"Error serving index: {e}", 500

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files"""
    return send_from_directory('static', filename)

# =========================
# Test Setup Routes
# =========================

@app.route('/api/test/setup', methods=['POST'])
def setup_test():
    """Setup a new test with answer keys"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['testTitle', 'staffName', 'totalQuestions', 'totalPages', 'marksPerQuestion', 'passingMarks']
        for field in required_fields:
            if field not in data:
                return jsonify({"success": False, "error": f"Missing required field: {field}"}), 400
        
        # Validate data types and ranges
        try:
            total_questions = int(data["totalQuestions"])
            total_pages = int(data["totalPages"])
            marks_per_question = float(data["marksPerQuestion"])
            passing_marks = float(data["passingMarks"])
        except ValueError:
            return jsonify({"success": False, "error": "Invalid numeric values"}), 400
        
        if total_questions <= 0 or total_pages <= 0:
            return jsonify({"success": False, "error": "Questions and pages must be positive numbers"}), 400
        
        if marks_per_question <= 0:
            return jsonify({"success": False, "error": "Marks per question must be positive"}), 400
        
        if passing_marks < 0 or passing_marks > 100:
            return jsonify({"success": False, "error": "Passing marks must be between 0-100"}), 400
        
        # Create test configuration
        test_config = {
            "id": datetime.now().strftime("%Y%m%d_%H%M%S"),
            "testTitle": data["testTitle"].strip(),
            "staffName": data["staffName"].strip(),
            "totalStudents": int(data.get("totalStudents", 50)),
            "totalQuestions": total_questions,
            "totalPages": total_pages,
            "marksPerQuestion": marks_per_question,
            "passingMarks": passing_marks,
            "answerKey": data.get("answerKey", []),
            "createdAt": datetime.now().isoformat(),
            "isActive": True
        }
        
        # Validate answer key
        if len(test_config["answerKey"]) != total_questions:
            return jsonify({"success": False, "error": "Answer key length must match total questions"}), 400
        
        # Save to answer keys file
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        
        # Deactivate previous tests
        for key in answer_keys:
            key["isActive"] = False
        
        # Add new test
        answer_keys.append(test_config)
        
        if save_json_data(ANSWER_KEYS_FILE, answer_keys):
            log_system_event("info", f"New test setup: {data['testTitle']} by {data['staffName']} - {total_questions} questions")
            return jsonify({"success": True, "testId": test_config["id"]})
        else:
            return jsonify({"success": False, "error": "Failed to save test configuration"}), 500
            
    except Exception as e:
        log_system_event("error", f"Test setup failed: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/test/active', methods=['GET'])
def get_active_test():
    """Get the currently active test"""
    try:
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        active_test = next((test for test in answer_keys if test.get("isActive", False)), None)
        
        if active_test:
            return jsonify({"success": True, "test": active_test})
        else:
            return jsonify({"success": False, "error": "No active test found"}), 404
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Student Evaluation Routes
# =========================

@app.route('/api/student/evaluate', methods=['POST'])
def evaluate_student():
    """Evaluate a student's answer sheet"""
    try:
        # Get active test
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        active_test = next((test for test in answer_keys if test.get("isActive", False)), None)
        
        if not active_test:
            return jsonify({"success": False, "error": "No active test found. Please set up a test first."}), 400
        
        student_name = request.form.get('studentName', '').strip()
        roll_number = request.form.get('rollNumber', '').strip()
        
        if not student_name or not roll_number:
            return jsonify({"success": False, "error": "Student name and roll number are required"}), 400
        
        # Process uploaded images
        predicted_answers = []
        image_paths = []
        
        # Save uploaded files and collect paths
        for i in range(1, active_test["totalPages"] + 1):
            file_key = f'answerSheet_{i}'
            if file_key not in request.files:
                return jsonify({"success": False, "error": f"Missing answer sheet for page {i}"}), 400
            
            file = request.files[file_key]
            if file.filename == '':
                return jsonify({"success": False, "error": f"No file selected for page {i}"}), 400
            
            if not allowed_file(file.filename):
                return jsonify({"success": False, "error": f"Invalid file type for page {i}. Allowed: PNG, JPG, JPEG, PDF"}), 400
            
            # Save file
            filename = secure_filename(f"{roll_number}_page_{i}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(filepath)
            image_paths.append(filepath)
        
        if len(image_paths) != active_test["totalPages"]:
            return jsonify({"success": False, "error": f"Expected {active_test['totalPages']} pages, got {len(image_paths)}"}), 400
        
        # Calculate questions per page
        total_questions = active_test["totalQuestions"]
        total_pages = active_test["totalPages"]
        questions_per_page = total_questions // total_pages
        remaining_questions = total_questions % total_pages
        
        log_system_event("info", f"Starting evaluation for {student_name} ({roll_number}) - {total_questions} questions across {total_pages} pages")
        
        # Process each page with AI
        for idx, image_path in enumerate(image_paths):
            page_num = idx + 1
            
            # Calculate questions for this page
            questions_on_page = questions_per_page
            if idx < remaining_questions:  # Distribute remaining questions to first pages
                questions_on_page += 1
            
            try:
                # Process image with AI
                page_answers = process_image_with_ai(image_path, questions_on_page)
                predicted_answers.extend(page_answers)
                
                log_system_event("info", f"Page {page_num} processed: {len(page_answers)} answers extracted")
                
            except Exception as e:
                log_system_event("error", f"Error processing page {page_num}: {str(e)}")
                # Add default answers for this page if processing fails
                predicted_answers.extend(['X'] * questions_on_page)
        
        # Ensure we have answers for all questions
        while len(predicted_answers) < total_questions:
            predicted_answers.append('X')
        
        # Truncate if we have too many answers
        predicted_answers = predicted_answers[:total_questions]
        
        # Compare with answer key
        answer_key = active_test["answerKey"]
        results = []
        correct_count = 0
        
        for i in range(total_questions):
            predicted = predicted_answers[i] if i < len(predicted_answers) else 'X'
            correct = answer_key[i] if i < len(answer_key) else 'A'
            
            is_correct = predicted == correct and predicted != 'X'
            if is_correct:
                correct_count += 1
            
            results.append({
                "questionNo": i + 1,
                "studentAnswer": predicted,
                "correctAnswer": correct,
                "isCorrect": is_correct,
                "marks": active_test["marksPerQuestion"] if is_correct else 0
            })
        
        # Calculate scores
        total_marks = total_questions * active_test["marksPerQuestion"]
        obtained_marks = correct_count * active_test["marksPerQuestion"]
        percentage = round((obtained_marks / total_marks) * 100, 2) if total_marks > 0 else 0
        
        # Determine grade and status
        if percentage >= 90:
            grade = 'A+'
        elif percentage >= 80:
            grade = 'A'
        elif percentage >= 70:
            grade = 'B+'
        elif percentage >= 60:
            grade = 'B'
        elif percentage >= 50:
            grade = 'C'
        elif percentage >= active_test["passingMarks"]:
            grade = 'D'
        else:
            grade = 'F'
        
        status = "PASS" if percentage >= active_test["passingMarks"] else "FAIL"
        
        # Create evaluation result
        evaluation_result = {
            "id": f"{roll_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            "studentName": student_name,
            "rollNumber": roll_number,
            "testTitle": active_test["testTitle"],
            "staffName": active_test["staffName"],
            "testId": active_test["id"],
            "results": results,
            "summary": {
                "totalQuestions": total_questions,
                "correctAnswers": correct_count,
                "wrongAnswers": total_questions - correct_count,
                "obtainedMarks": obtained_marks,
                "totalMarks": total_marks,
                "percentage": percentage,
                "grade": grade,
                "status": status
            },
            "evaluatedAt": datetime.now().isoformat(),
            "imagePaths": image_paths
        }
        
        log_system_event("info", f"Evaluation completed: {student_name} ({roll_number}) - {percentage:.1f}% ({status})")
        
        return jsonify({
            "success": True, 
            "evaluation": evaluation_result
        })
        
    except Exception as e:
        error_msg = str(e)
        log_system_event("error", f"Evaluation failed: {error_msg}")
        print(f"Full error traceback: {traceback.format_exc()}")
        return jsonify({"success": False, "error": error_msg}), 500

# =========================
# Consolidated Reports Routes
# =========================

@app.route('/api/reports/consolidated/save', methods=['POST'])
def save_consolidated_report():
    """Save a consolidated report after completing test evaluation"""
    try:
        data = request.get_json()
        
        if not data or 'consolidatedReport' not in data:
            return jsonify({"success": False, "error": "Invalid consolidated report data"}), 400
        
        consolidated_report = data['consolidatedReport']
        
        # Validate required fields
        required_fields = ['testId', 'testTitle', 'staffName', 'studentResults']
        for field in required_fields:
            if field not in consolidated_report:
                return jsonify({"success": False, "error": f"Missing required field: {field}"}), 400
        
        # Load existing consolidated reports
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        # Check if report already exists and update it
        existing_index = -1
        for i, report in enumerate(consolidated_reports):
            if report.get('testId') == consolidated_report['testId']:
                existing_index = i
                break
        
        # Add timestamps
        consolidated_report['savedAt'] = datetime.now().isoformat()
        if 'completedAt' not in consolidated_report:
            consolidated_report['completedAt'] = datetime.now().isoformat()
        
        if existing_index >= 0:
            # Update existing report
            consolidated_reports[existing_index] = consolidated_report
        else:
            # Add new report
            consolidated_reports.append(consolidated_report)
        
        if save_json_data(CONSOLIDATED_REPORTS_FILE, consolidated_reports):
            log_system_event("info", f"Consolidated report saved: {consolidated_report['testTitle']} - {len(consolidated_report['studentResults'])} students")
            return jsonify({"success": True, "reportId": consolidated_report['testId']})
        else:
            return jsonify({"success": False, "error": "Failed to save consolidated report"}), 500
            
    except Exception as e:
        log_system_event("error", f"Consolidated report save failed: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated', methods=['GET'])
def get_consolidated_reports():
    """Get all consolidated reports with statistics"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        # Sort by completion date (newest first)
        consolidated_reports.sort(key=lambda x: x.get('completedAt', '1970-01-01T00:00:00'), reverse=True)
        
        # Calculate overall statistics
        total_tests = len(consolidated_reports)
        total_students_evaluated = sum(len(report.get('studentResults', [])) for report in consolidated_reports)
        
        if total_tests > 0:
            overall_average = sum(report.get('averageScore', 0) for report in consolidated_reports) / total_tests
        else:
            overall_average = 0
        
        # Calculate additional statistics
        total_passed = sum(report.get('passedCount', 0) for report in consolidated_reports)
        total_failed = total_students_evaluated - total_passed
        overall_pass_rate = (total_passed / total_students_evaluated * 100) if total_students_evaluated > 0 else 0
        
        # Calculate highest and lowest scores across all reports
        all_scores = []
        for report in consolidated_reports:
            if report.get('studentResults'):
                for result in report['studentResults']:
                    all_scores.append(result.get('summary', {}).get('percentage', 0))
        
        highest_score = max(all_scores) if all_scores else 0
        lowest_score = min(all_scores) if all_scores else 0
        
        stats = {
            "totalTests": total_tests,
            "totalStudentsEvaluated": total_students_evaluated,
            "totalPassed": total_passed,
            "totalFailed": total_failed,
            "overallAverage": round(overall_average, 2),
            "overallPassRate": round(overall_pass_rate, 2),
            "highestScore": highest_score,
            "lowestScore": lowest_score
        }
        
        return jsonify({
            "success": True,
            "consolidatedReports": consolidated_reports,
            "stats": stats
        })
        
    except Exception as e:
        log_system_event("error", f"Failed to load consolidated reports: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated/<test_id>/export', methods=['GET'])
def export_consolidated_report(test_id):
    """Export a specific consolidated report to Excel"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        # Find the specific report
        report = next((r for r in consolidated_reports if r.get('testId') == test_id), None)
        if not report:
            return jsonify({"success": False, "error": "Consolidated report not found"}), 404
        
        # Create enhanced Excel workbook
        workbook = create_enhanced_excel_report(report)
        
        # Save to temporary file
        filename = f"consolidated_report_{report.get('testTitle', 'test').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, filename)
        
        workbook.save(filepath)
        
        log_system_event("info", f"Consolidated report exported: {report.get('testTitle')} - {filename}")
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_system_event("error", f"Consolidated report export failed: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated/export-custom', methods=['POST'])
def export_consolidated_report_custom():
    """Export consolidated report with custom Excel format from frontend"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400
        
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        # Create main sheet with student results
        ws = workbook.create_sheet("Student Results", 0)
        
        # Add headers
        headers = data.get('headers', [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        student_data = data.get('data', [])
        for row_idx, row_data in enumerate(student_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add statistics section
        statistics = data.get('statistics', [])
        if statistics:
            start_row = len(student_data) + 3
            for row_idx, stat_row in enumerate(statistics, start_row):
                for col_idx, value in enumerate(stat_row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1 and value:  # Bold first column
                        cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        test_info = data.get('testInfo', {})
        filename = f"{test_info.get('testTitle', 'exam').replace(' ', '_')}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, filename)
        
        workbook.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated/export-all', methods=['GET'])
def export_all_consolidated_reports():
    """Export all consolidated reports to a single Excel file"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        if not consolidated_reports:
            return jsonify({"success": False, "error": "No consolidated reports to export"}), 400
        
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        # Create overview sheet
        overview_sheet = workbook.create_sheet("Overview", 0)
        
        # Overview headers
        overview_headers = ['Test ID', 'Test Title', 'Staff Name', 'Total Students', 'Average Score (%)', 'Pass Rate (%)', 'Completed At']
        for col_idx, header in enumerate(overview_headers, 1):
            cell = overview_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Overview data
        for row_idx, report in enumerate(consolidated_reports, 2):
            completed_at = datetime.fromisoformat(report.get('completedAt', '')).strftime('%Y-%m-%d %H:%M') if report.get('completedAt') else ''
            row_data = [
                report.get('testId', ''),
                report.get('testTitle', ''),
                report.get('staffName', ''),
                report.get('totalStudents', 0),
                report.get('averageScore', 0),
                report.get('passRate', 0),
                completed_at
            ]
            for col_idx, value in enumerate(row_data, 1):
                overview_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths for overview
        for column in overview_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            overview_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create detailed sheet with all student results
        detailed_sheet = workbook.create_sheet("All Student Results", 1)
        
        # Detailed headers
        detailed_headers = [
            'Report ID', 'Roll Number', 'Student Name', 'Test Title', 'Staff Name',
            'Total Questions', 'Correct Answers', 'Wrong Answers', 'Obtained Marks',
            'Total Marks', 'Percentage (%)', 'Grade', 'Status', 'Evaluated At'
        ]
        
        for col_idx, header in enumerate(detailed_headers, 1):
            cell = detailed_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add all student data
        current_row = 2
        for report in consolidated_reports:
            for student in report.get('studentResults', []):
                summary = student.get('summary', {})
                evaluated_at = datetime.fromisoformat(
                    student.get('evaluatedAt', '').replace('Z', '+00:00')
                ).strftime('%Y-%m-%d %H:%M') if student.get('evaluatedAt') else ''
                
                row_data = [
                    report.get('testId', ''),
                    student.get('rollNumber', ''),
                    student.get('studentName', ''),
                    report.get('testTitle', ''),
                    report.get('staffName', ''),
                    summary.get('totalQuestions', 0),
                    summary.get('correctAnswers', 0),
                    summary.get('wrongAnswers', 0),
                    summary.get('obtainedMarks', 0),
                    summary.get('totalMarks', 0),
                    summary.get('percentage', 0),
                    summary.get('grade', ''),
                    summary.get('status', ''),
                    evaluated_at
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = detailed_sheet.cell(row=current_row, column=col_idx, value=value)
                    # Color code status
                    if col_idx == 13:  # Status column
                        if value == 'PASS':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                current_row += 1
        
        # Auto-adjust column widths for detailed sheet
        for column in detailed_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            detailed_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        filename = f"all_consolidated_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, filename)
        
        workbook.save(filepath)
        
        log_system_event("info", f"All consolidated reports exported: {filename}")
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_system_event("error", f"All consolidated reports export failed: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated/export-all-custom', methods=['GET'])
def export_all_consolidated_reports_custom():
    """Export all consolidated reports with enhanced formatting"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        if not consolidated_reports:
            return jsonify({"success": False, "error": "No consolidated reports to export"}), 400
        
        # Create comprehensive workbook
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        # 1. Executive Summary Sheet
        summary_sheet = workbook.create_sheet("Executive Summary", 0)
        
        # Calculate overall statistics
        total_tests = len(consolidated_reports)
        total_students = sum(len(report.get('studentResults', [])) for report in consolidated_reports)
        total_passed = sum(report.get('passedCount', 0) for report in consolidated_reports)
        total_failed = total_students - total_passed
        
        all_scores = []
        for report in consolidated_reports:
            if report.get('studentResults'):
                for result in report['studentResults']:
                    all_scores.append(result.get('summary', {}).get('percentage', 0))
        
        overall_average = sum(all_scores) / len(all_scores) if all_scores else 0
        highest_score = max(all_scores) if all_scores else 0
        lowest_score = min(all_scores) if all_scores else 0
        pass_rate = (total_passed / total_students * 100) if total_students > 0 else 0
        
        # Summary data
        summary_data = [
            ["SMART EXAM EVALUATOR - EXECUTIVE SUMMARY", ""],
            ["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["", ""],
            ["OVERALL STATISTICS", ""],
            ["Total Tests Conducted", total_tests],
            ["Total Students Evaluated", total_students],
            ["Students Passed", total_passed],
            ["Students Failed", total_failed],
            ["Overall Pass Rate (%)", f"{pass_rate:.1f}%"],
            ["Overall Average Score (%)", f"{overall_average:.1f}%"],
            ["Highest Score (%)", f"{highest_score:.1f}%"],
            ["Lowest Score (%)", f"{lowest_score:.1f}%"],
            ["", ""],
            ["PERFORMANCE ANALYSIS", ""],
            ["Excellent Performance (90%+)", sum(1 for score in all_scores if score >= 90)],
            ["Good Performance (70-89%)", sum(1 for score in all_scores if 70 <= score < 90)],
            ["Average Performance (50-69%)", sum(1 for score in all_scores if 50 <= score < 70)],
            ["Below Average (<50%)", sum(1 for score in all_scores if score < 50)],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            cell_a = summary_sheet.cell(row=row_idx, column=1, value=label)
            cell_b = summary_sheet.cell(row=row_idx, column=2, value=value)
            
            if "SUMMARY" in label or "STATISTICS" in label or "ANALYSIS" in label:
                cell_a.font = Font(bold=True, size=14)
                cell_a.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell_a.font = Font(bold=True, size=14, color="FFFFFF")
            elif label and value:
                cell_a.font = Font(bold=True)
        
        # Auto-adjust columns
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 20
        
        # 2. Test Overview Sheet
        overview_sheet = workbook.create_sheet("Test Overview", 1)
        
        headers = ['Test ID', 'Test Title', 'Staff Name', 'Date', 'Students', 'Avg Score', 'Pass Rate', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = overview_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, report in enumerate(consolidated_reports, 2):
            completed_at = datetime.fromisoformat(report.get('completedAt', '')).strftime('%Y-%m-%d') if report.get('completedAt') else ''
            status = "Completed" if report.get('studentResults') else "In Progress"
            
            row_data = [
                report.get('testId', ''),
                report.get('testTitle', ''),
                report.get('staffName', ''),
                completed_at,
                len(report.get('studentResults', [])),
                f"{report.get('averageScore', 0):.1f}%",
                f"{report.get('passRate', 0):.1f}%",
                status
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                overview_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust columns
        for column in overview_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            overview_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 3. Complete Student Results Sheet
        results_sheet = workbook.create_sheet("Complete Results", 2)
        
        result_headers = [
            'S.No', 'Test ID', 'Roll Number', 'Student Name', 'Test Title', 'Staff Name',
            'Questions', 'Correct', 'Wrong', 'Marks', 'Total', 'Percentage', 'Grade', 'Status', 'Date'
        ]
        
        for col_idx, header in enumerate(result_headers, 1):
            cell = results_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        current_row = 2
        serial_no = 1
        
        for report in consolidated_reports:
            for student in report.get('studentResults', []):
                summary = student.get('summary', {})
                evaluated_at = datetime.fromisoformat(
                    student.get('evaluatedAt', '').replace('Z', '+00:00')
                ).strftime('%Y-%m-%d') if student.get('evaluatedAt') else ''
                
                row_data = [
                    serial_no,
                    report.get('testId', ''),
                    student.get('rollNumber', ''),
                    student.get('studentName', ''),
                    report.get('testTitle', ''),
                    report.get('staffName', ''),
                    summary.get('totalQuestions', 0),
                    summary.get('correctAnswers', 0),
                    summary.get('wrongAnswers', 0),
                    summary.get('obtainedMarks', 0),
                    summary.get('totalMarks', 0),
                    f"{summary.get('percentage', 0):.1f}%",
                    summary.get('grade', ''),
                    summary.get('status', ''),
                    evaluated_at
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = results_sheet.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Color coding
                    if col_idx == 14:  # Status column
                        if value == 'PASS':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif col_idx == 13:  # Grade column
                        if value in ['A+', 'A']:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value in ['F']:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                current_row += 1
                serial_no += 1
        
        # Auto-adjust columns
        for column in results_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            results_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        filename = f"comprehensive_exam_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, filename)
        
        workbook.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/reports/consolidated/<test_id>', methods=['DELETE'])
def delete_consolidated_report(test_id):
    """Delete a specific consolidated report"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        # Find and remove the report
        initial_count = len(consolidated_reports)
        consolidated_reports = [r for r in consolidated_reports if r.get('testId') != test_id]
        
        if len(consolidated_reports) == initial_count:
            return jsonify({"success": False, "error": "Consolidated report not found"}), 404
        
        if save_json_data(CONSOLIDATED_REPORTS_FILE, consolidated_reports):
            log_system_event("info", f"Consolidated report deleted: {test_id}")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to delete consolidated report"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Admin Routes
# =========================

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Admin login authentication"""
    try:
        data = request.get_json()
        password = data.get('password', '')
        
        # Load settings to get admin password
        settings = load_json_data(SETTINGS_FILE, {"adminPassword": "admin123"})
        admin_password = settings.get("adminPassword", "admin123")
        
        if password == admin_password:
            log_system_event("info", "Admin login successful")
            return jsonify({"success": True})
        else:
            log_system_event("warning", "Admin login failed - incorrect password")
            return jsonify({"success": False, "error": "Incorrect password"}), 401
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/stats', methods=['GET'])
def get_admin_stats():
    """Get system statistics"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        
        # Calculate statistics
        total_evaluations = sum(len(report.get('studentResults', [])) for report in consolidated_reports)
        active_tests = len([test for test in answer_keys if test.get("isActive", False)])
        
        # Get last backup time
        backup_files = []
        try:
            backup_files = [f for f in os.listdir(app.config["UPLOAD_FOLDER"]) 
                          if f.startswith("system_backup_") or f.startswith("consolidated_report_")]
        except:
            pass
            
        last_backup = "Never"
        if backup_files:
            backup_files.sort(reverse=True)
            try:
                latest_backup = backup_files[0]
                if "system_backup_" in latest_backup:
                    date_str = latest_backup.replace("system_backup_", "").replace(".json", "")
                elif "consolidated_report_" in latest_backup:
                    date_str = latest_backup.split("_")[-2] + "_" + latest_backup.split("_")[-1].replace(".xlsx", "")
                backup_date = datetime.strptime(date_str, "%Y%m%d_%H%M%S")
                last_backup = backup_date.strftime("%Y-%m-%d %H:%M")
            except:
                last_backup = "Available"
        
        stats = {
            "totalEvaluations": total_evaluations,
            "activeTests": active_tests,
            "systemUptime": get_system_uptime(),
            "lastBackup": last_backup
        }
        
        return jsonify({"success": True, "stats": stats})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/logs', methods=['GET'])
def get_system_logs():
    """Get system logs"""
    try:
        logs = load_json_data(LOGS_FILE, [])
        # Return last 100 logs
        logs = logs[:100]
        return jsonify({"success": True, "logs": logs})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/clear-logs', methods=['POST'])
def clear_logs():
    """Clear system logs"""
    try:
        if save_json_data(LOGS_FILE, []):
            log_system_event("info", "System logs cleared by admin")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to clear logs"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/delete-consolidated-reports', methods=['POST'])
def delete_all_consolidated_reports():
    """Delete all consolidated reports"""
    try:
        if save_json_data(CONSOLIDATED_REPORTS_FILE, []):
            log_system_event("warning", "All consolidated reports deleted by admin")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to delete consolidated reports"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/clear-answer-keys', methods=['POST'])
def clear_answer_keys():
    """Clear answer keys"""
    try:
        if save_json_data(ANSWER_KEYS_FILE, []):
            log_system_event("warning", "Answer keys cleared by admin")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to clear answer keys"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/backup', methods=['POST'])
def create_backup():
    """Create system backup"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create backup data
        backup_data = {
            "timestamp": datetime.now().isoformat(),
            "consolidatedReports": load_json_data(CONSOLIDATED_REPORTS_FILE, []),
            "answer_keys": load_json_data(ANSWER_KEYS_FILE, []),
            "settings": load_json_data(SETTINGS_FILE, {}),
            "logs": load_json_data(LOGS_FILE, [])
        }
        
        # Save backup file
        backup_filename = f"system_backup_{timestamp}.json"
        backup_path = os.path.join(app.config["UPLOAD_FOLDER"], backup_filename)
        
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, indent=2, ensure_ascii=False)
        
        log_system_event("info", f"System backup created: {backup_filename}")
        
        return jsonify({
            "success": True, 
            "backupFile": backup_filename,
            "timestamp": timestamp
        })
        
    except Exception as e:
        log_system_event("error", f"Backup creation failed: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/settings', methods=['GET', 'POST'])
def admin_settings():
    """Get or update admin settings"""
    try:
        if request.method == 'GET':
            settings = load_json_data(SETTINGS_FILE, {
                "refreshInterval": 30,
                "soundAlerts": True,
                "autoSave": True,
                "adminPassword": "admin123"
            })
            # Don't return the password
            safe_settings = {k: v for k, v in settings.items() if k != 'adminPassword'}
            return jsonify({"success": True, "settings": safe_settings})
        
        elif request.method == 'POST':
            data = request.get_json()
            current_settings = load_json_data(SETTINGS_FILE, {})
            
            # Update settings
            for key, value in data.items():
                if key != 'adminPassword':  # Don't allow password change via this endpoint
                    current_settings[key] = value
            
            if save_json_data(SETTINGS_FILE, current_settings):
                log_system_event("info", "Admin settings updated")
                return jsonify({"success": True})
            else:
                return jsonify({"success": False, "error": "Failed to save settings"}), 500
                
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin/change-password', methods=['POST'])
def change_admin_password():
    """Change admin password"""
    try:
        data = request.get_json()
        current_password = data.get('currentPassword', '')
        new_password = data.get('newPassword', '')
        
        if not current_password or not new_password:
            return jsonify({"success": False, "error": "Both current and new passwords are required"}), 400
        
        # Load current settings
        settings = load_json_data(SETTINGS_FILE, {"adminPassword": "admin123"})
        
        # Verify current password
        if current_password != settings.get("adminPassword", "admin123"):
            return jsonify({"success": False, "error": "Current password is incorrect"}), 401
        
        # Update password
        settings["adminPassword"] = new_password
        
        if save_json_data(SETTINGS_FILE, settings):
            log_system_event("info", "Admin password changed successfully")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to save new password"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Dashboard Stats Routes
# =========================

@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        total_students = sum(len(report.get('studentResults', [])) for report in consolidated_reports)
        
        if total_students > 0:
            # Calculate average score across all students
            all_percentages = []
            latest_date = None
            
            for report in consolidated_reports:
                for student in report.get('studentResults', []):
                    summary = student.get('summary', {})
                    if 'percentage' in summary:
                        all_percentages.append(summary['percentage'])
                
                # Track latest evaluation date
                completion_date = report.get('completedAt', '1970-01-01T00:00:00')
                if not latest_date or completion_date > latest_date:
                    latest_date = completion_date
            
            avg_score = sum(all_percentages) / len(all_percentages) if all_percentages else 0
            
            # Format the date nicely
            last_eval = "Never"
            if latest_date:
                try:
                    last_eval_date = datetime.fromisoformat(latest_date.replace('Z', '+00:00'))
                    last_eval = last_eval_date.strftime("%Y-%m-%d %H:%M")
                except:
                    last_eval = "Unknown"
        else:
            avg_score = 0
            last_eval = "Never"
        
        # Get active test info
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        active_test = next((test for test in answer_keys if test.get("isActive", False)), None)
        active_test_name = active_test.get("testTitle", "None") if active_test else "None"
        
        stats = {
            "totalStudents": total_students,
            "avgScore": f"{avg_score:.1f}%",
            "lastEval": last_eval,
            "activeTest": active_test_name,
            "totalTests": len(consolidated_reports)
        }
        
        return jsonify({"success": True, "stats": stats})
        
    except Exception as e:
        log_system_event("error", f"Failed to load dashboard stats: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# File Management Routes
# =========================

@app.route('/api/files/cleanup', methods=['POST'])
def cleanup_files():
    """Clean up old uploaded files"""
    try:
        upload_folder = app.config["UPLOAD_FOLDER"]
        cutoff_days = request.json.get('days', 7) if request.json else 7
        cutoff_date = datetime.now() - timedelta(days=cutoff_days)
        
        deleted_count = 0
        total_size_freed = 0
        
        if os.path.exists(upload_folder):
            for filename in os.listdir(upload_folder):
                filepath = os.path.join(upload_folder, filename)
                
                if os.path.isfile(filepath):
                    # Check file modification time
                    file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                    
                    if file_time < cutoff_date and not filename.startswith('system_backup_'):
                        try:
                            file_size = os.path.getsize(filepath)
                            os.remove(filepath)
                            deleted_count += 1
                            total_size_freed += file_size
                        except Exception as e:
                            print(f"Error deleting {filepath}: {e}")
        
        # Convert bytes to MB
        size_mb = total_size_freed / (1024 * 1024)
        
        log_system_event("info", f"File cleanup completed: {deleted_count} files deleted, {size_mb:.1f}MB freed")
        
        return jsonify({
            "success": True,
            "deletedFiles": deleted_count,
            "sizeFreed": f"{size_mb:.1f}MB"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/files/stats', methods=['GET'])
def get_file_stats():
    """Get file storage statistics"""
    try:
        upload_folder = app.config["UPLOAD_FOLDER"]
        
        total_files = 0
        total_size = 0
        file_types = {}
        
        if os.path.exists(upload_folder):
            for filename in os.listdir(upload_folder):
                filepath = os.path.join(upload_folder, filename)
                
                if os.path.isfile(filepath):
                    total_files += 1
                    file_size = os.path.getsize(filepath)
                    total_size += file_size
                    
                    # Count file types
                    ext = filename.split('.')[-1].lower() if '.' in filename else 'unknown'
                    file_types[ext] = file_types.get(ext, 0) + 1
        
        # Convert bytes to MB
        size_mb = total_size / (1024 * 1024)
        
        return jsonify({
            "success": True,
            "stats": {
                "totalFiles": total_files,
                "totalSize": f"{size_mb:.1f}MB",
                "fileTypes": file_types
            }
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Health Check Routes
# =========================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Check AI model availability
        ai_status = "available" if model else "unavailable"
        
        # Check disk space
        try:
            disk_usage = os.statvfs('.')
            free_space_mb = (disk_usage.f_bavail * disk_usage.f_frsize) / (1024 * 1024)
        except:
            free_space_mb = 0
        
        # Check data files
        files_status = {
            "consolidatedReports": os.path.exists(CONSOLIDATED_REPORTS_FILE),
            "answerKeys": os.path.exists(ANSWER_KEYS_FILE),
            "settings": os.path.exists(SETTINGS_FILE),
            "logs": os.path.exists(LOGS_FILE)
        }
        
        health_data = {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "uptime": get_system_uptime(),
            "aiModel": ai_status,
            "diskSpace": f"{free_space_mb:.1f}MB",
            "dataFiles": files_status
        }
        
        return jsonify({"success": True, "health": health_data})
        
    except Exception as e:
        return jsonify({
            "success": False, 
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }), 500

# =========================
# Enhanced API Routes for Frontend Compatibility
# =========================

@app.route('/api/test/list', methods=['GET'])
def list_tests():
    """Get list of all tests (active and inactive)"""
    try:
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        
        # Sort by creation date (newest first)
        answer_keys.sort(key=lambda x: x.get('createdAt', ''), reverse=True)
        
        return jsonify({"success": True, "tests": answer_keys})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/test/<test_id>/activate', methods=['POST'])
def activate_test(test_id):
    """Activate a specific test"""
    try:
        answer_keys = load_json_data(ANSWER_KEYS_FILE, [])
        
        # Deactivate all tests first
        for test in answer_keys:
            test["isActive"] = False
        
        # Find and activate the specified test
        test_found = False
        for test in answer_keys:
            if test.get("id") == test_id:
                test["isActive"] = True
                test_found = True
                break
        
        if not test_found:
            return jsonify({"success": False, "error": "Test not found"}), 404
        
        if save_json_data(ANSWER_KEYS_FILE, answer_keys):
            log_system_event("info", f"Test activated: {test_id}")
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to activate test"}), 500
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/student/bulk-upload', methods=['POST'])
def bulk_upload_students():
    """Handle bulk student evaluation from uploaded CSV/Excel file"""
    try:
        if 'studentFile' not in request.files:
            return jsonify({"success": False, "error": "No student file uploaded"}), 400
        
        file = request.files['studentFile']
        if file.filename == '':
            return jsonify({"success": False, "error": "No file selected"}), 400
        
        # Process the uploaded file
        # This would typically involve reading CSV/Excel and processing each student
        # For now, return a placeholder response
        
        return jsonify({
            "success": True,
            "message": "Bulk upload feature is under development",
            "studentsProcessed": 0
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/analytics/performance', methods=['GET'])
def get_performance_analytics():
    """Get detailed performance analytics"""
    try:
        consolidated_reports = load_json_data(CONSOLIDATED_REPORTS_FILE, [])
        
        if not consolidated_reports:
            return jsonify({"success": True, "analytics": {}})
        
        # Calculate comprehensive analytics
        analytics = {
            "overallStats": {},
            "testComparisons": [],
            "performanceTrends": [],
            "subjectAnalysis": {},
            "gradeDistribution": {}
        }
        
        # Overall statistics
        total_students = sum(len(report.get('studentResults', [])) for report in consolidated_reports)
        all_scores = []
        grade_counts = {'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        
        for report in consolidated_reports:
            for student in report.get('studentResults', []):
                summary = student.get('summary', {})
                all_scores.append(summary.get('percentage', 0))
                grade = summary.get('grade', 'F')
                if grade in grade_counts:
                    grade_counts[grade] += 1
        
        if all_scores:
            analytics["overallStats"] = {
                "totalStudents": total_students,
                "averageScore": sum(all_scores) / len(all_scores),
                "highestScore": max(all_scores),
                "lowestScore": min(all_scores),
                "standardDeviation": np.std(all_scores) if len(all_scores) > 1 else 0
            }
        
        analytics["gradeDistribution"] = grade_counts
        
        # Test comparisons
        for report in consolidated_reports:
            test_scores = []
            for student in report.get('studentResults', []):
                test_scores.append(student.get('summary', {}).get('percentage', 0))
            
            if test_scores:
                analytics["testComparisons"].append({
                    "testId": report.get('testId'),
                    "testTitle": report.get('testTitle'),
                    "averageScore": sum(test_scores) / len(test_scores),
                    "passRate": report.get('passRate', 0),
                    "studentCount": len(test_scores)
                })
        
        return jsonify({"success": True, "analytics": analytics})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Error Handlers
# =========================

@app.errorhandler(404)
def not_found(error):
    return jsonify({
        "success": False, 
        "error": "Endpoint not found",
        "message": "The requested API endpoint does not exist"
    }), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        "success": False, 
        "error": "Internal server error",
        "message": "An unexpected error occurred on the server"
    }), 500

@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({
        "success": False,
        "error": "File too large",
        "message": "The uploaded file exceeds the maximum size limit"
    }), 413

@app.errorhandler(400)
def bad_request(error):
    return jsonify({
        "success": False,
        "error": "Bad request",
        "message": "The request was malformed or invalid"
    }), 400

# =========================
# Initialize System
# =========================

def initialize_system():
    """Initialize system on startup"""
    print("🚀 Initializing Smart Exam Evaluator System v2.2.0...")
    
    # Log system start
    log_system_event("info", "Smart Exam Evaluator system started")
    
    # Create default settings if not exists
    default_settings = {
        "refreshInterval": 30,
        "soundAlerts": True,
        "autoSave": True,
        "adminPassword": "admin123",
        "maxFileSize": 50,  # MB
        "allowedFileTypes": ["png", "jpg", "jpeg", "pdf", "bmp", "tiff"],
        "autoCleanupDays": 30,
        "systemName": "Smart Exam Evaluator",
        "version": "2.2.0"
    }
    
    if not os.path.exists(SETTINGS_FILE):
        save_json_data(SETTINGS_FILE, default_settings)
        print("✅ Default settings created")
    
    # Initialize empty data files if they don't exist
    if not os.path.exists(CONSOLIDATED_REPORTS_FILE):
        save_json_data(CONSOLIDATED_REPORTS_FILE, [])
        print("✅ Consolidated reports file initialized")
    
    if not os.path.exists(ANSWER_KEYS_FILE):
        save_json_data(ANSWER_KEYS_FILE, [])
        print("✅ Answer keys file initialized")
    
    if not os.path.exists(LOGS_FILE):
        save_json_data(LOGS_FILE, [])
        print("✅ Logs file initialized")
    
    # Check AI model
    if model:
        print("✅ Gemini AI model ready")
    else:
        print("⚠️  Gemini AI model not available - check API key")
    
    # Print system info
    print(f"📁 Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"📁 Data folder: {app.config['DATA_FOLDER']}")
    print(f"🔧 Max file size: {app.config['MAX_CONTENT_LENGTH'] / (1024*1024):.0f}MB")
    
    print("✅ System initialization complete!")
    print("=" * 50)

# =========================
# Development Helper Routes
# =========================

@app.route('/api/dev/reset', methods=['POST'])
def reset_system():
    """Reset system data (development only)"""
    if not app.debug:
        return jsonify({"success": False, "error": "Not available in production"}), 403
    
    try:
        # Clear all data files
        save_json_data(CONSOLIDATED_REPORTS_FILE, [])
        save_json_data(ANSWER_KEYS_FILE, [])
        save_json_data(LOGS_FILE, [])
        
        # Clear uploads folder (except backups)
        upload_folder = app.config["UPLOAD_FOLDER"]
        if os.path.exists(upload_folder):
            for filename in os.listdir(upload_folder):
                if not filename.startswith('system_backup_'):
                    try:
                        os.remove(os.path.join(upload_folder, filename))
                    except:
                        pass
        
        log_system_event("warning", "System reset by developer")
        return jsonify({"success": True, "message": "System reset complete"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/dev/sample-data', methods=['POST'])
def create_sample_data():
    """Create sample data for testing (development only)"""
    if not app.debug:
        return jsonify({"success": False, "error": "Not available in production"}), 403
    
    try:
        # Create sample test
        sample_test = {
            "id": "sample_test_001",
            "testTitle": "Sample Mathematics Test",
            "staffName": "Dr. John Smith",
            "totalStudents": 30,
            "totalQuestions": 10,
            "totalPages": 2,
            "marksPerQuestion": 1.0,
            "passingMarks": 40.0,
            "answerKey": ["A", "B", "C", "D", "A", "B", "C", "D", "A", "B"],
            "createdAt": datetime.now().isoformat(),
            "isActive": True
        }
        
        answer_keys = [sample_test]
        save_json_data(ANSWER_KEYS_FILE, answer_keys)
        
        # Create sample consolidated report
        student_names = ["Alice Johnson", "Bob Smith", "Carol Davis", "David Wilson", "Eva Brown"]
        roll_numbers = ["2023001", "2023002", "2023003", "2023004", "2023005"]
        
        student_results = []
        total_percentage = 0
        passed_count = 0
        
        for i, (name, roll) in enumerate(zip(student_names, roll_numbers)):
            correct_answers = 7 + i % 3  # Vary scores
            percentage = (correct_answers / 10) * 100
            total_percentage += percentage
            
            if percentage >= 40:  # Passing marks
                passed_count += 1
            
            student_result = {
                "id": f"sample_{roll}",
                "studentName": name,
                "rollNumber": roll,
                "testTitle": "Sample Mathematics Test",
                "staffName": "Dr. John Smith",
                "testId": "sample_test_001",
                "results": [
                    {"questionNo": j+1, "studentAnswer": "A", "correctAnswer": "A", "isCorrect": j < correct_answers, "marks": 1 if j < correct_answers else 0}
                    for j in range(10)
                ],
                "summary": {
                    "totalQuestions": 10,
                    "correctAnswers": correct_answers,
                    "wrongAnswers": 10 - correct_answers,
                    "obtainedMarks": correct_answers,
                    "totalMarks": 10,
                    "percentage": percentage,
                    "grade": "A" if percentage >= 80 else "B" if percentage >= 60 else "C" if percentage >= 40 else "F",
                    "status": "PASS" if percentage >= 40 else "FAIL"
                },
                "evaluatedAt": (datetime.now() - timedelta(days=i)).isoformat(),
                "imagePaths": []
            }
            student_results.append(student_result)
        
        # Create consolidated report
        avg_score = total_percentage / len(student_results)
        pass_rate = (passed_count / len(student_results)) * 100
        highest_score = max(student['summary']['percentage'] for student in student_results)
        lowest_score = min(student['summary']['percentage'] for student in student_results)
        
        consolidated_report = {
            "testId": "sample_test_001",
            "testTitle": "Sample Mathematics Test",
            "staffName": "Dr. John Smith",
            "totalStudents": 5,
            "totalQuestions": 10,
            "marksPerQuestion": 1.0,
            "passingMarks": 40.0,
            "completedAt": datetime.now().isoformat(),
            "studentResults": student_results,
            "averageScore": round(avg_score, 2),
            "highestScore": highest_score,
            "lowestScore": lowest_score,
            "passRate": round(pass_rate, 2),
            "passedCount": passed_count,
            "failedCount": len(student_results) - passed_count,
            "savedAt": datetime.now().isoformat()
        }
        
        consolidated_reports = [consolidated_report]
        save_json_data(CONSOLIDATED_REPORTS_FILE, consolidated_reports)
        
        log_system_event("info", "Sample data created for development")
        return jsonify({"success": True, "message": "Sample data created"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# Main Application
# =========================

if __name__ == '__main__':
    initialize_system()
    
    # Run the app
    print("🌐 Starting Flask server...")
    print("📱 Access the application at: http://localhost:5000")
    print("🔧 Admin panel available after login")
    print("📊 Enhanced Excel export functionality enabled")
    print("🤖 AI-powered answer sheet processing ready")
    print("=" * 50)
    
    app.run(
        debug=True, 
        host='0.0.0.0', 
        port=5000,
        threaded=True
    )
